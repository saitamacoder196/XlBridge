"""Feature 2: Inject translated content back into Excel.

Supported entry types (determined by the ``coord`` field in each TXT line)
---------------------------------------------------------------------------
cell  — [SheetName]!A1|...          → write directly to cell
shape — [SheetName]!shape:Name|...  → write to shape text body by name
note  — [SheetName]!note:A1|...     → write to cell comment/note

Supported languages for the ``lang`` parameter
-----------------------------------------------
None  — use the original value (column 1); backward-compatible default.
"en"  — use the English translation (column 2).
        Falls back to the original value if the EN column is absent.
"vi"  — use the Vietnamese translation (column 3).
        Falls back to the original value if the VI column is absent.

Any other language string raises ``ValueError`` (not supported).
"""

import logging
import shutil
import warnings
from pathlib import Path
from typing import Literal

import openpyxl
from openpyxl.comments import Comment

from xlbridge.parser import CellEntry, parse_txt

logger = logging.getLogger(__name__)

# XML namespaces for drawing XML (shape injection)
_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
_A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'

Lang = Literal['en', 'vi'] | None
SUPPORTED_LANGS: set[str] = {'en', 'vi'}


# ── value picker ──────────────────────────────────────────────────────────────

def _pick_value(entry: CellEntry, lang: Lang) -> tuple[str, bool]:
    """Select the value for the requested language.

    Returns:
        (value, used_fallback) — used_fallback is True when the requested
        translation column was absent and the original text was used instead.
    """
    if lang == 'en':
        if entry.en is not None:
            return entry.en, False
        logger.debug('EN translation missing for %s!%s — using original', entry.sheet, entry.coord)
        return entry.value, True

    if lang == 'vi':
        if entry.vi is not None:
            return entry.vi, False
        logger.debug('VI translation missing for %s!%s — using original', entry.sheet, entry.coord)
        return entry.value, True

    # lang is None → original value (backward-compatible)
    return entry.value, False


# ── shape injection ───────────────────────────────────────────────────────────

def _inject_shape(ws, shape_name: str, value: str) -> bool:
    """Write ``value`` into the named shape's text body.

    Finds the first ``<xdr:sp>`` element whose ``cNvPr name`` matches
    ``shape_name``, then replaces all text content with ``value`` while
    keeping the first run's formatting attributes intact.

    Returns True if the shape was found and updated, False otherwise.
    """
    drawings = getattr(ws, '_drawings', [])
    for drawing in drawings:
        elem = getattr(drawing, '_element', None)
        if elem is None:
            continue
        try:
            for sp in elem.iter(f'{{{_XDR}}}sp'):
                cNvPr = sp.find(f'{{{_XDR}}}nvSpPr/{{{_XDR}}}cNvPr')
                if cNvPr is None or cNvPr.get('name') != shape_name:
                    continue

                txBody = sp.find(f'{{{_XDR}}}txBody')
                if txBody is None:
                    continue

                # Collect every <a:r> run across all paragraphs
                all_runs = list(txBody.iter(f'{{{_A}}}r'))
                if not all_runs:
                    continue

                # Set the first run's <a:t> to the new text, remove the rest
                first_t = all_runs[0].find(f'{{{_A}}}t')
                if first_t is not None:
                    first_t.text = value
                for run in all_runs[1:]:
                    parent = run.getparent()
                    if parent is not None:
                        parent.remove(run)

                # Keep only the first paragraph; remove extras
                paragraphs = txBody.findall(f'{{{_A}}}p')
                for p in paragraphs[1:]:
                    txBody.remove(p)

                return True

        except Exception as exc:  # pylint: disable=broad-except
            logger.debug("Shape injection error for '%s': %s", shape_name, exc)

    return False


# ── note injection ────────────────────────────────────────────────────────────

def _inject_note(ws, cell_coord: str, value: str) -> None:
    """Update (or create) the note/comment on ``cell_coord`` with ``value``."""
    try:
        cell = ws[cell_coord]
        author = cell.comment.author if cell.comment else ''
        cell.comment = Comment(value, author)
    except Exception as exc:  # pylint: disable=broad-except
        logger.warning('Failed to inject note at %s: %s', cell_coord, exc)


# ── public API ────────────────────────────────────────────────────────────────

def inject(
    input_path: str,
    translation_path: str,
    output_path: str | None = None,
    lang: Lang = None,
) -> str:
    """Inject text from a TXT file into an Excel file.

    Handles cells, shapes, and notes in the same pass.
    Creates a new file; never overwrites the original.

    Args:
        input_path:       Path to the original .xlsx file.
        translation_path: Path to the TXT file (2- or 4-column format).
        output_path:      Path for the output file.
                          Defaults to ``*_<lang>.xlsx`` when lang is set,
                          otherwise ``*_translated.xlsx``.
        lang:             Target language — ``"en"``, ``"vi"``, or ``None``.

    Returns:
        The path of the output file.

    Raises:
        ValueError: If ``lang`` is not None, ``"en"``, or ``"vi"``.
    """
    if lang is not None and lang not in SUPPORTED_LANGS:
        raise ValueError(
            f"Language '{lang}' is not supported. "
            f"Supported values: {sorted(SUPPORTED_LANGS)} or None."
        )

    if output_path is None:
        p = Path(input_path)
        suffix = f'_{lang}' if lang else '_translated'
        output_path = str(p.with_name(f'{p.stem}{suffix}{p.suffix}'))

    # Copy original so all formatting / formulas are preserved
    shutil.copy2(input_path, output_path)

    entries = parse_txt(translation_path)
    with warnings.catch_warnings():
        warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
        wb = openpyxl.load_workbook(output_path)

    applied       = 0
    skipped_sheet = 0
    skipped_shape = 0
    fallback      = 0

    for entry in entries:
        if entry.sheet not in wb.sheetnames:
            logger.warning(
                "Sheet '%s' not found, skipping %s!%s",
                entry.sheet, entry.sheet, entry.coord,
            )
            skipped_sheet += 1
            continue

        value, used_fallback = _pick_value(entry, lang)
        if used_fallback:
            fallback += 1

        ws = wb[entry.sheet]

        if entry.entry_type == 'cell':
            ws[entry.coord] = value
            applied += 1

        elif entry.entry_type == 'shape':
            if _inject_shape(ws, entry.shape_name, value):
                applied += 1
            else:
                logger.warning(
                    "Shape '%s' not found in sheet '%s'",
                    entry.shape_name, entry.sheet,
                )
                skipped_shape += 1

        elif entry.entry_type == 'note':
            _inject_note(ws, entry.note_coord, value)
            applied += 1

    wb.save(output_path)
    wb.close()

    logger.info(
        'Injected %d/%d entries into %s  '
        '(lang=%s, fallback=%d, sheet_missing=%d, shape_missing=%d)',
        applied, len(entries), output_path,
        lang or 'original', fallback, skipped_sheet, skipped_shape,
    )
    return output_path
