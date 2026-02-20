"""Helper functions for XlBridge."""

from openpyxl.utils import get_column_letter


def cell_coord(col: int, row: int) -> str:
    """Convert 1-based column and row to cell coordinate like 'A1'."""
    return f"{get_column_letter(col)}{row}"


def is_merged_slave(sheet, col: int, row: int) -> bool:
    """Check if a cell is part of a merged range but NOT the top-left origin."""
    for merged_range in sheet.merged_cells.ranges:
        if (merged_range.min_row <= row <= merged_range.max_row
                and merged_range.min_col <= col <= merged_range.max_col):
            if row != merged_range.min_row or col != merged_range.min_col:
                return True
    return False
