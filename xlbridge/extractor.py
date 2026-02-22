"""Feature 1: Extract Excel cell content to TXT.

Output format
-------------
Each non-empty cell, shape, and note is written as one line:

    [SheetName]!A1|value               ← cell
    [SheetName]!shape:TextBox 1|text   ← shape / text-box
    [SheetName]!note:A1|comment text   ← cell comment / note

All three forms use the same pipe-separated layout and are fully parseable
by the XlBridge parser.  Newlines inside values are escaped as ``\\n``.
"""

import logging
from datetime import date
from pathlib import Path

import openpyxl

from xlbridge.utils import cell_coord, is_merged_slave

logger = logging.getLogger(__name__)

# XML namespaces used in Excel drawing XML (SpreadsheetML drawing)
_XDR = 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing'
_A   = 'http://schemas.openxmlformats.org/drawingml/2006/main'


# ── per-sheet extractors ──────────────────────────────────────────────────────

def _extract_cells(ws, sheet_name: str) -> list[str]:
    """Extract non-empty cell values from a worksheet."""
    lines: list[str] = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if is_merged_slave(ws, col, row):
                continue
            cell = ws.cell(row=row, column=col)
            if cell.value is None or str(cell.value).strip() == '':
                continue
            coord = cell_coord(col, row)
            value = str(cell.value).replace('\n', '\\n')
            lines.append(f'[{sheet_name}]!{coord}|{value}')
    return lines


def _extract_shapes(ws, sheet_name: str) -> list[str]:
    """Extract text from shapes (text boxes, auto shapes) via the drawing XML.

    Uses openpyxl's internal ``_drawings`` attribute (stable across 3.x) and
    parses the raw lxml element tree directly for reliability.
    """
    lines: list[str] = []
    drawings = getattr(ws, '_drawings', [])
    for drawing in drawings:
        elem = getattr(drawing, '_element', None)
        if elem is None:
            continue
        try:
            for sp in elem.iter(f'{{{_XDR}}}sp'):
                # Shape name from <xdr:cNvPr name="..." id="..."/>
                cNvPr = sp.find(f'{{{_XDR}}}nvSpPr/{{{_XDR}}}cNvPr')
                shape_name = ''
                if cNvPr is not None:
                    shape_name = (cNvPr.get('name') or '').strip()
                    if not shape_name:
                        shape_name = str(cNvPr.get('id', '?'))
                if not shape_name:
                    shape_name = '?'

                # Collect all <a:t> text runs inside the shape
                texts = [t.text or '' for t in sp.iter(f'{{{_A}}}t')]
                text = ''.join(texts).strip()
                if not text:
                    continue

                value = text.replace('\n', '\\n')
                lines.append(f'[{sheet_name}]!shape:{shape_name}|{value}')

        except Exception as exc:  # pylint: disable=broad-except
            logger.debug("Shape extraction error in sheet '%s': %s", sheet_name, exc)

    return lines


def _extract_notes(ws, sheet_name: str) -> list[str]:
    """Extract cell comments/notes from a worksheet."""
    lines: list[str] = []
    for row in ws.iter_rows():
        for cell in row:
            comment = cell.comment
            if comment is None:
                continue
            text = (comment.text or '').strip()
            if not text:
                continue
            value = text.replace('\n', '\\n')
            lines.append(f'[{sheet_name}]!note:{cell.coordinate}|{value}')
    return lines


# ── public API ────────────────────────────────────────────────────────────────

def extract(
    input_path: str,
    output_path: str,
    sheet_names: list[str] | None = None,
    include_shapes: bool = True,
    include_notes: bool = True,
) -> None:
    """Extract cells, shapes, and notes from an Excel file to a TXT file.

    Args:
        input_path:     Path to the source .xlsx file.
        output_path:    Path for the output .txt file.
        sheet_names:    Sheet names to extract. ``None`` means all sheets.
        include_shapes: Include text from shapes / text boxes (default: True).
        include_notes:  Include cell comments / notes (default: True).
    """
    wb = openpyxl.load_workbook(input_path, data_only=True)
    source_name = Path(input_path).name
    lines: list[str] = []

    total_cells  = 0
    total_shapes = 0
    total_notes  = 0

    sheets = sheet_names if sheet_names else wb.sheetnames
    for name in sheets:
        if name not in wb.sheetnames:
            logger.warning("Sheet '%s' not found, skipping.", name)
            continue

        ws = wb[name]

        # ── cells ──
        cell_lines = _extract_cells(ws, name)
        total_cells += len(cell_lines)
        lines.extend(cell_lines)

        # ── shapes ──
        if include_shapes:
            shape_lines = _extract_shapes(ws, name)
            total_shapes += len(shape_lines)
            if shape_lines:
                lines.append(f'# -- shapes: {name} --')
                lines.extend(shape_lines)

        # ── notes ──
        if include_notes:
            note_lines = _extract_notes(ws, name)
            total_notes += len(note_lines)
            if note_lines:
                lines.append(f'# -- notes: {name} --')
                lines.extend(note_lines)

    wb.close()

    extras = []
    if include_shapes:
        extras.append(f'Shapes: {total_shapes}')
    if include_notes:
        extras.append(f'Notes: {total_notes}')
    extras_str = ('  ' + '  '.join(extras)) if extras else ''

    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('# XlBridge Export\n')
        f.write(f'# Source: {source_name}\n')
        f.write(f'# Date: {date.today().isoformat()}\n')
        f.write(f'# Encoding: UTF-8\n')
        f.write(f'# Cells: {total_cells}{extras_str}\n')
        f.write('\n')
        for line in lines:
            f.write(line + '\n')

    logger.info(
        "Extracted %d cells, %d shapes, %d notes → %s",
        total_cells, total_shapes, total_notes, output_path,
    )
