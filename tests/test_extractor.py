"""Tests for the extractor."""

import openpyxl

from xlbridge.extractor import extract
from xlbridge.parser import parse_txt


def _create_test_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Hello"
    ws["B1"] = "World"
    ws["A2"] = "日本語"

    ws2 = wb.create_sheet("Sheet2")
    ws2["A1"] = "Data"

    wb.save(path)
    wb.close()


def test_extract_all_sheets(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt = tmp_path / "out.txt"
    _create_test_xlsx(str(xlsx))

    extract(str(xlsx), str(txt))
    entries = parse_txt(str(txt))

    sheets = {e.sheet for e in entries}
    assert "Sheet1" in sheets
    assert "Sheet2" in sheets
    assert len(entries) == 4


def test_extract_single_sheet(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt = tmp_path / "out.txt"
    _create_test_xlsx(str(xlsx))

    extract(str(xlsx), str(txt), sheet_names=["Sheet1"])
    entries = parse_txt(str(txt))

    sheets = {e.sheet for e in entries}
    assert sheets == {"Sheet1"}
    assert len(entries) == 3


def test_extract_merged_cells(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt = tmp_path / "out.txt"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Merged"
    ws.merge_cells("A1:B1")
    ws["A2"] = "Normal"
    wb.save(str(xlsx))
    wb.close()

    extract(str(xlsx), str(txt))
    entries = parse_txt(str(txt))

    coords = [e.coord for e in entries]
    assert "A1" in coords
    assert "B1" not in coords
    assert "A2" in coords
