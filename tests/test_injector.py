"""Tests for the injector."""

import pytest
import openpyxl

from xlbridge.injector import inject


def _create_test_xlsx(path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Original"
    ws["B1"] = "Keep"
    ws["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    wb.save(str(path))
    wb.close()


def _write_txt(path, content):
    path.write_text(content, encoding="utf-8")


# ─── Backward-compatible (2-column, no lang) ──────────────────────────────────

def test_inject_basic(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[Sheet1]!A1|Translated\n")

    output = inject(str(xlsx), str(txt))
    wb = openpyxl.load_workbook(output)

    assert wb["Sheet1"]["A1"].value == "Translated"
    assert wb["Sheet1"]["B1"].value == "Keep"
    wb.close()


def test_inject_preserves_format(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[Sheet1]!A1|New\n")

    output = inject(str(xlsx), str(txt))
    wb = openpyxl.load_workbook(output)

    assert wb["Sheet1"]["A1"].value == "New"
    assert wb["Sheet1"]["A1"].font.bold is True
    assert wb["Sheet1"]["A1"].font.size == 14
    wb.close()


def test_inject_default_output_name_no_lang(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[Sheet1]!A1|Test\n")

    output = inject(str(xlsx), str(txt))
    assert "_translated" in output


def test_inject_missing_sheet_skipped(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[NoSheet]!A1|Test\n")

    output = inject(str(xlsx), str(txt))
    wb = openpyxl.load_workbook(output)
    assert wb["Sheet1"]["A1"].value == "Original"
    wb.close()


# ─── 4-column format with lang="en" ──────────────────────────────────────────

def test_inject_lang_en(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[Sheet1]!A1|原文|English value|Giá trị tiếng Việt\n")

    output = inject(str(xlsx), str(txt), lang="en")
    wb = openpyxl.load_workbook(output)

    assert wb["Sheet1"]["A1"].value == "English value"
    wb.close()


def test_inject_lang_vi(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[Sheet1]!A1|原文|English value|Giá trị tiếng Việt\n")

    output = inject(str(xlsx), str(txt), lang="vi")
    wb = openpyxl.load_workbook(output)

    assert wb["Sheet1"]["A1"].value == "Giá trị tiếng Việt"
    wb.close()


def test_inject_lang_en_fallback_to_original(tmp_path):
    """When EN column is absent, fall back to original value."""
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[Sheet1]!A1|原文のみ\n")   # 2-column, no EN

    output = inject(str(xlsx), str(txt), lang="en")
    wb = openpyxl.load_workbook(output)

    assert wb["Sheet1"]["A1"].value == "原文のみ"
    wb.close()


def test_inject_lang_vi_fallback_to_original(tmp_path):
    """When VI column is absent, fall back to original value."""
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[Sheet1]!A1|Original|English only\n")  # 3-column, no VI

    output = inject(str(xlsx), str(txt), lang="vi")
    wb = openpyxl.load_workbook(output)

    assert wb["Sheet1"]["A1"].value == "Original"
    wb.close()


def test_inject_default_output_name_with_lang_en(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[Sheet1]!A1|原文|English|Việt\n")

    output = inject(str(xlsx), str(txt), lang="en")
    assert output.endswith("_en.xlsx")


def test_inject_default_output_name_with_lang_vi(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[Sheet1]!A1|原文|English|Việt\n")

    output = inject(str(xlsx), str(txt), lang="vi")
    assert output.endswith("_vi.xlsx")


# ─── Unsupported language ─────────────────────────────────────────────────────

def test_inject_unsupported_lang_raises(tmp_path):
    xlsx = tmp_path / "test.xlsx"
    txt  = tmp_path / "trans.txt"
    _create_test_xlsx(xlsx)
    _write_txt(txt, "[Sheet1]!A1|原文|English|Việt\n")

    with pytest.raises(ValueError, match="not supported"):
        inject(str(xlsx), str(txt), lang="ja")
